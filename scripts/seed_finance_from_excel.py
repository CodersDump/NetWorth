"""
NetWorth - seed the finance table from the Badminton_court_expenses.xlsx
workbook's "Calculations" sheet (the normalized layer: expenses J-Q,
memberships Z-AD, walk-ins A-H). This is the STARTING data load - all
future entries happen through the site's Finance tab, not this script.

Player linking rules (as confirmed):
    - walk-in "Aditya Yelegaona"        -> roster "Adi"
    - "Aditya Nair"                     -> roster "Aditya"
    - "Sohan Kuchangari"                -> roster "Sohan" (pays per-session)
    - "Prasanna"                        -> roster "Prasanna" (his own account;
                                           Sambvit is a DIFFERENT person)
    - "Saurabh Tiwari"                  -> roster roster "Saurabh T"
    - "Bibhudatta"/"Vibhudatta"         -> roster "Bibhu"
    - "Sandeep Rathore"/"Sandeep"       -> roster "Sandeep"
    - "Sashi"/"Shashi", "Udit"          -> roster (joined monthly after walk-in)
    - walk-in "Mohit" is a DIFFERENT person than roster Mohit - never linked
    - "Ram Shukla" not yet registered - stays a name string until he is
    - anything else: linked only on an exact name match, else kept as text

Usage:
    python seed_finance_from_excel.py --xlsx Badminton_court_expenses.xlsx           (dry run: prints everything, writes nothing)
    python seed_finance_from_excel.py --xlsx Badminton_court_expenses.xlsx --apply   (writes to DynamoDB)

Safe to re-run: record_ids are deterministic (uuid5 of the record content
key), so re-running overwrites the same records instead of duplicating.
"""
import argparse
import re
import uuid
from decimal import Decimal

import boto3
import openpyxl

NAMESPACE = uuid.UUID('a5b1c3d4-0000-4000-8000-networth0000'.replace('networth', '12ab34cd'))

EXCEL_TO_SITE = {
    'Aditya Yelegaona': 'Adi',
    'Aditya Nair': 'Aditya',
    'Sohan Kuchangari': 'Sohan',
    'Bibhudatta': 'Bibhu',
    'Vibhudatta': 'Bibhu',
    'Sandeep Rathore': 'Sandeep',
    'Sashi': 'Shashi',
    'Saurabh Tiwari': 'Saurabh T',
}
# NOTE: 'Prasanna' intentionally has NO mapping - he is registered under his
# own name; 'Sambvit' is a different person entirely.
WALKIN_NEVER_LINK = {'Mohit'}  # different person than the roster Mohit

MONTHS = ['January', 'February', 'March', 'April', 'May', 'June', 'July',
          'August', 'September', 'October', 'November', 'December']


def det_id(*parts):
    return str(uuid.uuid5(NAMESPACE, '|'.join(str(p) for p in parts)))


def parse_day(day_str):
    m = re.match(r'(\d+)', str(day_str))
    return int(m.group(1)) if m else None


def iso_date(day_str, month_name, year):
    day = parse_day(day_str)
    if day is None or month_name not in MONTHS:
        return None
    return f"{int(year):04d}-{MONTHS.index(month_name) + 1:02d}-{day:02d}"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--xlsx', required=True)
    parser.add_argument('--apply', action='store_true', help='write to DynamoDB (default: dry run)')
    parser.add_argument('--region', default='us-east-1')
    args = parser.parse_args()

    # read_only computes the sheet's real extent; the normal loader trusts
    # the declared dimension (1,048,576 rows here) and iter_rows would then
    # walk a million empty rows. Belt and suspenders: hard cap as well.
    wb = openpyxl.load_workbook(args.xlsx, data_only=True, read_only=True)
    ws = wb['Calculations']
    max_row = min(ws.max_row or 0, 10000)

    dynamodb = boto3.resource('dynamodb', region_name=args.region)
    finance_table = dynamodb.Table('networth-finance')
    players_table = dynamodb.Table('networth-players')

    players = players_table.scan().get('Items', [])
    name_to_pid = {p['name']: p['player_id'] for p in players}

    def norm(n):
        return ' '.join(str(n).lower().split())
    norm_to_player = {norm(p['name']): p for p in players}

    linked, unlinked = set(), set()

    def resolve(excel_name, is_walkin=False):
        excel_name = str(excel_name).strip()
        if is_walkin and excel_name in WALKIN_NEVER_LINK:
            unlinked.add(f"{excel_name} (walk-in; intentionally NOT linked to roster {excel_name})")
            return excel_name, None
        site_name = EXCEL_TO_SITE.get(excel_name, excel_name)
        pid = name_to_pid.get(site_name)
        if pid:
            linked.add(f"{excel_name} -> {site_name}")
            return site_name, pid
        # Fallback: case/whitespace-insensitive match against the roster,
        # so a registration typed as "abhishek v" still links "Abhishek V".
        p = norm_to_player.get(norm(site_name))
        if p:
            linked.add(f"{excel_name} -> {p['name']} (case-insensitive match)")
            return p['name'], p['player_id']
        unlinked.add(excel_name)
        return excel_name, None

    records = []

    # ---- expenses: J..Q, header row 1 ----
    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=10, max_col=17):
        item, month, year, slot, est_cost, act_cost, est_qty, act_qty = (c.value for c in row)
        if not item or not month:
            continue
        records.append({
            'record_id': det_id('expense', month, year, slot, item),
            'record_type': 'expense',
            'month': str(month), 'year': Decimal(str(int(year))), 'slot': str(slot),
            'item': str(item),
            'estimated_cost': Decimal(str(est_cost)), 'actual_cost': Decimal(str(act_cost)),
            'estimated_qty': Decimal(str(est_qty)), 'actual_qty': Decimal(str(act_qty)),
        })

    # ---- memberships: Z..AD, header row 1 ----
    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=26, max_col=30):
        name, slot, month, year, status = (c.value for c in row)
        if not name or not month:
            continue
        display, pid = resolve(name)
        rec = {
            'record_id': det_id('membership', name, month, year, slot),
            'record_type': 'membership',
            'month': str(month), 'year': Decimal(str(int(year))), 'slot': str(slot),
            'display_name': display, 'status': str(status or 'No'),
        }
        if pid:
            rec['player_id'] = pid
        records.append(rec)

    # ---- walk-ins: A..H, header row 1 ----
    seq = {}
    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=8):
        name, day, month, year, slot, fee, skill, recruit = (c.value for c in row)
        if not name or not month:
            continue
        date = iso_date(day, str(month), year)
        if not date:
            print(f"SKIPPED walk-in row (bad date): {name} {day} {month} {year}")
            continue
        display, pid = resolve(name, is_walkin=True)
        # Same person can appear twice on one date (e.g. paid both slots or a
        # correction row) - a sequence number keeps deterministic ids unique.
        key = (name, date, slot)
        seq[key] = seq.get(key, 0) + 1
        rec = {
            'record_id': det_id('walkin', name, date, slot, seq[key]),
            'record_type': 'walkin',
            'date': date, 'slot': str(slot), 'display_name': display,
            'fee': Decimal(str(fee if fee is not None else 0)),
        }
        if pid:
            rec['player_id'] = pid
        if skill and str(skill) != 'NA':
            rec['skill'] = str(skill)
        if recruit and str(recruit) != 'NA':
            rec['recruit_verdict'] = str(recruit)
        if rec['fee'] < 0:
            rec['note'] = 'refund/adjustment (imported from Excel as a negative fee)'
        records.append(rec)

    by_type = {}
    for r in records:
        by_type.setdefault(r['record_type'], []).append(r)
    print(f"Parsed: {len(by_type.get('expense', []))} expenses, "
          f"{len(by_type.get('membership', []))} memberships, "
          f"{len(by_type.get('walkin', []))} walk-ins\n")
    print("LINKED to roster players:")
    for l in sorted(linked):
        print(f"  {l}")
    print("\nNOT linked (kept as name strings - link later via the Finance tab if they register):")
    for u in sorted(unlinked):
        print(f"  {u}")

    if not args.apply:
        print("\nDRY RUN - nothing written. Re-run with --apply to seed DynamoDB.")
        return

    with finance_table.batch_writer() as batch:
        for r in records:
            batch.put_item(Item=r)
    print(f"\nSeeded {len(records)} records into networth-finance.")


if __name__ == '__main__':
    main()
